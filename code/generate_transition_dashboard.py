from __future__ import annotations

import json
import math
import shutil
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
XLS_CANDIDATES = [
    ROOT / "outputs" / "mip_epe_replication_results.xlsx",
    ROOT / "data" / "raw" / "mip_epe_replication_results.xlsx",
]
OUT_HTML = ROOT / "outputs" / "mip_epe_dashboard.html"
DOCS_HTML = ROOT / "docs" / "index.html"


def read_sheet(wb, name: str) -> list[dict]:
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    out = []
    for row in rows[1:]:
        rec = {}
        empty = True
        for h, value in zip(headers, row):
            if not h:
                continue
            if value is not None:
                empty = False
            rec[h] = value
        if not empty:
            out.append(rec)
    return out


def num(x, default=0.0):
    if x is None:
        return default
    try:
        if isinstance(x, str) and not x.strip():
            return default
        v = float(x)
        if math.isnan(v):
            return default
        return v
    except Exception:
        return default


def pct(n, d):
    return None if not d else n / d * 100


def build_data(xls: Path) -> dict:
    wb = load_workbook(xls, read_only=True, data_only=True)
    sheets = {s: read_sheet(wb, s) for s in wb.sheetnames}

    setores = {r["cod"]: r for r in sheets.get("Setores", []) if r.get("cod")}
    baseline = {r["cod"]: r for r in sheets.get("Baseline", []) if r.get("cod")}
    energy = {r["cod"]: r for r in sheets.get("Energia_Fluxos_ktep", []) if r.get("cod")}

    if sheets.get("Satellite"):
        satellite = sheets["Satellite"]
    else:
        satellite = []
        for cod, b in baseline.items():
            e = energy.get(cod, {})
            s = setores.get(cod, {})
            output = num(b.get("vbp"))
            jobs = num(b.get("ocupacoes"))
            wages = num(b.get("remuner"))
            gdp = num(b.get("va_pib"))
            deriv = num(e.get("Derivados"))
            biod = num(e.get("Biodiesel"))
            etanol = num(e.get("Etanol"))
            eec = num(e.get("EE_Central"))
            eed = num(e.get("EE_Distrib"))
            gas = num(e.get("Gas_Natural"))
            total = deriv + biod + etanol + eec + eed + gas
            fossil = deriv + gas
            lowcarbon = biod + etanol + eec + eed
            satellite.append(
                {
                    "cod": cod,
                    "nome": b.get("nome") or s.get("nome"),
                    "grupo": s.get("grupo") or b.get("grupo"),
                    "grupo_nt4": s.get("grupo_nt4"),
                    "eh_energia": bool(s.get("eh_energia")),
                    "eh_fossil": bool(s.get("eh_fossil")),
                    "eh_renov": bool(s.get("eh_renov")),
                    "output": output,
                    "gdp": gdp,
                    "jobs": jobs,
                    "wages": wages,
                    "energy_ktep": total,
                    "fossil_ktep": fossil,
                    "lowcarbon_ktep": lowcarbon,
                    "labor_coef_raw": jobs / output if output else None,
                    "labor_coef_nt": 0.4 * jobs / output if output else None,
                    "wage_coef": wages / output if output else None,
                    "gdp_coef": gdp / output if output else None,
                    "energy_coef": total / output if output else None,
                    "fossil_energy_coef": fossil / output if output else None,
                    "lowcarbon_energy_coef": lowcarbon / output if output else None,
                    "fossil_share": fossil / total if total else None,
                    "lowcarbon_share": lowcarbon / total if total else None,
                }
            )

    shocks_sector = sheets.get("Choques_Setor", [])
    # If workbook still contains the pre-calibration employment shock, include NT-calibrated view.
    for r in shocks_sector:
        r["delta_emp_raw"] = num(r.get("delta_emp"))
        r["delta_emp_nt"] = num(r.get("delta_emp")) * 0.4
        if "delta_energy_total" not in r:
            sat = next((x for x in satellite if x.get("cod") == r.get("cod")), {})
            dx = num(r.get("delta_x"))
            r["delta_energy_total"] = num(sat.get("energy_coef")) * dx
            r["delta_fossil_energy"] = num(sat.get("fossil_energy_coef")) * dx
            r["delta_lowcarbon_energy"] = num(sat.get("lowcarbon_energy_coef")) * dx

    shock_summary = []
    for ex in sorted({r.get("exercicio") for r in shocks_sector if r.get("exercicio")}):
        rows = [r for r in shocks_sector if r.get("exercicio") == ex]
        shock_summary.append(
            {
                "exercicio": ex,
                "dx_bi": sum(num(r.get("delta_x")) for r in rows) / 1000,
                "dVA_bi": sum(num(r.get("delta_va")) for r in rows) / 1000,
                "demp_raw_mil": sum(num(r.get("delta_emp_raw")) for r in rows) / 1000,
                "demp_nt_mil": sum(num(r.get("delta_emp_nt")) for r in rows) / 1000,
                "dE_ktep": sum(num(r.get("delta_energy_total")) for r in rows),
            }
        )

    return {
        "source": str(xls.relative_to(ROOT)),
        "satellite": satellite,
        "shocks_sector": shocks_sector,
        "shocks_energy": sheets.get("Choques_Energia", []),
        "shocks_labor": sheets.get("Choques_Labor", []),
        "shock_summary": shock_summary,
        "has_native_satellite": bool(sheets.get("Satellite")),
    }


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>MIP-EPE Transition Dashboard</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.2/dist/chart.umd.min.js"></script>
<style>
:root{--bg:#0f1117;--panel:#181c25;--line:#293040;--text:#e7edf6;--muted:#93a0b4;--blue:#378ADD;--green:#1D9E75;--red:#E24B4A;--gold:#BA7517}
*{box-sizing:border-box}body{margin:0;background:var(--bg);color:var(--text);font-family:Inter,Segoe UI,Arial,sans-serif;font-size:14px}
header{padding:18px 28px;border-bottom:1px solid var(--line);background:#121620;position:sticky;top:0;z-index:10}
h1{font-size:20px;margin:0 0 4px}header p{margin:0;color:var(--muted);font-size:12px}
main{padding:22px 28px;max-width:1500px;margin:auto}.grid{display:grid;gap:16px}.g4{grid-template-columns:repeat(4,1fr)}.g3{grid-template-columns:repeat(3,1fr)}.g2{grid-template-columns:repeat(2,1fr)}
.card{background:var(--panel);border:1px solid var(--line);border-radius:9px;padding:16px}.title{font-size:12px;color:var(--muted);text-transform:uppercase;letter-spacing:.08em;margin-bottom:12px;font-weight:700}
.kpi .v{font-size:28px;font-weight:800}.kpi .l{color:var(--muted);font-size:12px;margin-top:3px}.chart{height:320px}.tall{height:430px}
.note{color:var(--muted);font-size:12px;line-height:1.45}.pill{display:inline-block;border:1px solid var(--line);border-radius:999px;padding:5px 10px;margin:2px;color:var(--muted)}
table{width:100%;border-collapse:collapse;font-size:12px}th,td{padding:7px 8px;border-bottom:1px solid #242b39;text-align:left}th{color:var(--muted);background:#1d2330;position:sticky;top:0}tr:hover td{background:#202737}
select,input{background:#1d2330;color:var(--text);border:1px solid var(--line);border-radius:6px;padding:7px 9px}
.toolbar{display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin-bottom:12px}.pos{color:var(--green)}.neg{color:var(--red)}
@media(max-width:900px){.g4,.g3,.g2{grid-template-columns:1fr}.chart{height:280px}}
</style>
</head>
<body>
<header>
  <h1>MIP-EPE Energy Transition Dashboard</h1>
  <p>Satellite-account indicators connected to Leontief shocks. Source workbook: <span id="src"></span></p>
</header>
<main class="grid">
  <section class="grid g4">
    <div class="card kpi"><div class="v" id="kOutput">-</div><div class="l">Baseline output, R$ trillion</div></div>
    <div class="card kpi"><div class="v" id="kJobs">-</div><div class="l">Baseline employment, million jobs</div></div>
    <div class="card kpi"><div class="v" id="kEnergy">-</div><div class="l">Baseline final energy, ktep</div></div>
    <div class="card kpi"><div class="v" id="kFossil">-</div><div class="l">Fossil share of final energy</div></div>
  </section>
  <section class="grid g2">
    <div class="card"><div class="title">Satellite Coefficients: Top Labor Intensity</div><div class="chart"><canvas id="laborChart"></canvas></div></div>
    <div class="card"><div class="title">Satellite Coefficients: Top Energy Intensity</div><div class="chart"><canvas id="energyChart"></canvas></div></div>
  </section>
  <section class="grid g2">
    <div class="card"><div class="title">Transition Exposure: Fossil Share x Energy Intensity</div><div class="chart tall"><canvas id="scatterChart"></canvas></div></div>
    <div class="card"><div class="title">Shock Summary: GDP, Employment, Energy</div><div class="chart tall"><canvas id="shockChart"></canvas></div></div>
  </section>
  <section class="grid g3">
    <div class="card"><div class="title">Employment Lens</div><div class="note" id="empNote"></div></div>
    <div class="card"><div class="title">Energy Lens</div><div class="note" id="energyNote"></div></div>
    <div class="card"><div class="title">Dashboard Notes</div><div class="note" id="dashNote"></div></div>
  </section>
  <section class="card">
    <div class="toolbar">
      <div class="title" style="margin:0">Sector Satellite Table</div>
      <input id="search" placeholder="Search sector/code..." oninput="renderTable()">
      <select id="sort" onchange="renderTable()">
        <option value="energy_coef">Sort by energy coefficient</option>
        <option value="labor_coef_nt">Sort by NT labor coefficient</option>
        <option value="fossil_share">Sort by fossil share</option>
        <option value="output">Sort by output</option>
      </select>
    </div>
    <div style="max-height:520px;overflow:auto">
      <table id="tbl"><thead><tr>
        <th>Code</th><th>Sector</th><th>Group</th><th>Output</th><th>Jobs</th>
        <th>Labor coef NT</th><th>Energy coef</th><th>Fossil share</th><th>Low-carbon share</th>
      </tr></thead><tbody></tbody></table>
    </div>
  </section>
</main>
<script>
const DATA = __DATA__;
const S = DATA.satellite || [];
const SH = DATA.shock_summary || [];
const fmt = (x,d=1)=>Number(x||0).toLocaleString("en-US",{maximumFractionDigits:d,minimumFractionDigits:d});
const sum = (a,f)=>a.reduce((s,r)=>s+(Number(r[f])||0),0);
const color = {blue:"#378ADD",green:"#1D9E75",red:"#E24B4A",gold:"#BA7517",gray:"#8892a4"};
document.getElementById("src").textContent = DATA.source;
document.getElementById("kOutput").textContent = fmt(sum(S,"output")/1e6,2);
document.getElementById("kJobs").textContent = fmt(sum(S,"jobs")/1e6,1);
document.getElementById("kEnergy").textContent = fmt(sum(S,"energy_ktep"),0);
document.getElementById("kFossil").textContent = fmt(sum(S,"fossil_ktep")/Math.max(sum(S,"energy_ktep"),1)*100,1)+"%";
const top = (field,n=15)=>[...S].filter(r=>isFinite(Number(r[field]))).sort((a,b)=>(b[field]||0)-(a[field]||0)).slice(0,n);
function bar(id, labels, datasets, horizontal=false){
  new Chart(document.getElementById(id),{type:"bar",data:{labels,datasets},options:{responsive:true,maintainAspectRatio:false,indexAxis:horizontal?"y":"x",plugins:{legend:{labels:{color:"#e7edf6"}}},scales:{x:{ticks:{color:"#93a0b4"},grid:{color:"#293040"}},y:{ticks:{color:"#93a0b4"},grid:{color:"#293040"}}}}});
}
bar("laborChart", top("labor_coef_nt").map(r=>r.cod), [{label:"jobs/R$M, NT calibrated",data:top("labor_coef_nt").map(r=>r.labor_coef_nt),backgroundColor:color.green}], true);
bar("energyChart", top("energy_coef").map(r=>r.cod), [{label:"ktep/R$M",data:top("energy_coef").map(r=>r.energy_coef),backgroundColor:color.gold}], true);
new Chart(document.getElementById("scatterChart"),{type:"scatter",data:{datasets:[{label:"Sectors",data:S.map(r=>({x:(r.fossil_share||0)*100,y:r.energy_coef||0,cod:r.cod,nome:r.nome,g:r.grupo})),backgroundColor:S.map(r=>r.eh_fossil?color.red:(r.eh_renov?color.green:color.blue)),pointRadius:S.map(r=>r.eh_energia?7:4)}]},options:{responsive:true,maintainAspectRatio:false,plugins:{tooltip:{callbacks:{label:c=>`${c.raw.cod} ${c.raw.nome}: fossil ${fmt(c.raw.x,1)}%, energy coef ${fmt(c.raw.y,4)}`}}},scales:{x:{title:{display:true,text:"Fossil share of sector energy (%)",color:"#93a0b4"},ticks:{color:"#93a0b4"},grid:{color:"#293040"}},y:{title:{display:true,text:"Energy coefficient (ktep/R$M)",color:"#93a0b4"},ticks:{color:"#93a0b4"},grid:{color:"#293040"}}}}});
bar("shockChart", SH.map(r=>r.exercicio.replace(" (NT 6.1)","").replace(" (NT 6.2)","")), [
  {label:"GDP impact, R$ bi",data:SH.map(r=>r.dVA_bi),backgroundColor:color.blue,yAxisID:"y"},
  {label:"Employment NT, thousand",data:SH.map(r=>r.demp_nt_mil),backgroundColor:color.green,yAxisID:"y"},
  {label:"Energy, ktep",data:SH.map(r=>r.dE_ktep),backgroundColor:color.gold,yAxisID:"y1"}
], false);
Chart.getChart("shockChart").options.scales.y1={position:"right",ticks:{color:"#93a0b4"},grid:{drawOnChartArea:false,color:"#293040"}}; Chart.getChart("shockChart").update();
const energyTop = top("energy_coef",3).map(r=>`<span class="pill">${r.cod}: ${fmt(r.energy_coef,4)} ktep/R$M</span>`).join("");
const laborTop = top("labor_coef_nt",3).map(r=>`<span class="pill">${r.cod}: ${fmt(r.labor_coef_nt,2)} jobs/R$M</span>`).join("");
document.getElementById("empNote").innerHTML = `Shock employment uses the NT-calibrated coefficient: <b>0.4 × occupations/output</b>. This preserves Anexo 2 baseline jobs while matching NT shock magnitudes.<br><br>${laborTop}`;
document.getElementById("energyNote").innerHTML = `Energy impacts by source come from ALPHA × Δx. Sector transition exposure is summarized by total energy intensity and fossil share.<br><br>${energyTop}`;
document.getElementById("dashNote").innerHTML = DATA.has_native_satellite ? "Native Satellite sheets found in workbook." : "Satellite indicators were computed inside this dashboard from Baseline and Energia_Fluxos_ktep. Rerun the R replication script to export gender, informality, education, and carbon columns directly.";
function renderTable(){
  const q=document.getElementById("search").value.toLowerCase();
  const sort=document.getElementById("sort").value;
  const rows=[...S].filter(r=>(r.cod+" "+r.nome+" "+r.grupo).toLowerCase().includes(q)).sort((a,b)=>(b[sort]||0)-(a[sort]||0));
  document.querySelector("#tbl tbody").innerHTML=rows.map(r=>`<tr>
    <td>${r.cod}</td><td>${r.nome}</td><td>${r.grupo||""}</td><td>${fmt((r.output||0)/1000,1)} bi</td>
    <td>${fmt(r.jobs||0,0)}</td><td>${fmt(r.labor_coef_nt,3)}</td><td>${fmt(r.energy_coef,5)}</td>
    <td>${r.fossil_share==null?"-":fmt(r.fossil_share*100,1)+"%"}</td><td>${r.lowcarbon_share==null?"-":fmt(r.lowcarbon_share*100,1)+"%"}</td>
  </tr>`).join("");
}
renderTable();
</script>
</body>
</html>
"""


def main() -> None:
    xls = next((p for p in XLS_CANDIDATES if p.exists()), None)
    if xls is None:
        raise FileNotFoundError("Could not find mip_epe_replication_results.xlsx")
    data = build_data(xls)
    html = HTML_TEMPLATE.replace("__DATA__", json.dumps(data, ensure_ascii=False, default=str))
    OUT_HTML.parent.mkdir(parents=True, exist_ok=True)
    OUT_HTML.write_text(html, encoding="utf-8")
    DOCS_HTML.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(OUT_HTML, DOCS_HTML)
    print(f"Wrote {OUT_HTML}")
    print(f"Updated {DOCS_HTML}")


if __name__ == "__main__":
    main()
